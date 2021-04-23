from openpyxl import load_workbook
from untils.readYaml import baseConfig
import xlrd
tiqu_path=baseConfig.TIQU_PATH
api_case_path=baseConfig.APICASE_EXCEL_PATH

class Excel_rw():
    def __init__(self,filename):
        self.filename=filename
        self.wb = load_workbook(self.filename)  # 读取一个excel
        sheets_name = self.wb.sheetnames  # 获取所有的sheet表格名称，返回的
        self.sheet1 = self.wb[sheets_name[0]]  # 第一个sheet表格
        '''xlrd'''
        self.excel=xlrd.open_workbook(filename)

    def get_col_value(self):
        '''获取第一列所有值，组装成一个列表'''
        rows = self.sheet1.max_row  # 4
        data=[]
        for i in range(2, rows + 1):  # i=2,3,4,excel列表第二行，开始读取
            clo_value = self.sheet1.cell(row=i, column=1).value
            data.append(clo_value)
        return data

    def write_tiqu(self,rowv,colV):
        #获取有多少行
        rows=self.sheet1.max_row
        #判断是否存在相同的变量名称 1.获取所有的变量名称，组成一个列表。2.新写入的变量是否在列表中
        V_names = self.get_col_value()#['message','code']
        if rowv in V_names:
            #更新值
            idx=V_names.index(rowv)#获取已存在的变量名称的索引+2
            self.sheet1.cell(row=idx+2,column=2).value=colV
        else:
            #写入第一列的值 ：变量名称
            self.sheet1.cell(row=rows+1,column=1).value=rowv
            #写入第二列的值 ：变量具体的值
            self.sheet1.cell(row=rows+1,column=2).value=colV
        self.wb.save(self.filename)

    def read_tiqu(self,name):
        #循环第一列的值
        rows=self.sheet1.max_row  #4
        for i in range(2,rows+1): #i=2,3,4
            clo_value=self.sheet1.cell(row=i, column=1).value
            if clo_value==name:
                data=self.sheet1.cell(row=i, column=2).value
                return data
        #如果值等于name，就取这行的第二列的值

    def read_api_cases(self):
        sheets=self.excel.sheet_names() #获取有多少个表格，返回的是列表
        # print('表格数量',sheets)
        casesInfo = []  # 所有用例的数据的汇总是列表，列表的元素为字典，字典代表每个用例的信息
        for s in range(len(sheets)):
            sheet = self.excel.sheet_by_index(s)  # 第几个sheet表格
            rows, cols = sheet.nrows, sheet.ncols  # 获取行数和列数
            headers = sheet.row_values(0)  # excel的头部,固定的
            # print(rows,cols,headers,)
            for i in range(1, rows):  # 每次循环之后，会往列表里面添加一个新的字典，也就是新的用例
                caseInfo = {}  # 字典赋值的格式：{key头部信息第一列，value对应每一行第一个值}
                for j in range(0, cols):  # 等有多少列全部循环之后，添加字典
                    caseInfo[headers[j]] = sheet.row_values(i)[j]
                casesInfo.append(caseInfo)
        #print(casesInfo)
        return casesInfo
    # ['接口用例名称', '接口url', '请求方式', '请求头信息', '请求入参', '提取变量', '检查点']
    # ['登录1',‘login’,'post','','']
    # ['登录2',‘login’,'post','',]

    #{"接口用例名称": "登录", "接口入参": "{}"}


    # 把excel sheet解析为双层列表，每一个行是一个外层元素，每个单元格是一个内层元素。pytest使用。
    def sheet_to_pytestlist(self):
        lsts = []
        sheets = self.excel.sheet_names()
        #添加excel表格头的信息，pytest参数化的参数名
        lsts.append(["host,casename,apiname,method,headers,data,extract,duanyan,run"])
        # nrows是sheet的行数
        for s in range(len(sheets)):
            sheet = self.excel.sheet_by_index(s)  # 第几个sheet表格
            rows, cols = sheet.nrows, sheet.ncols  # 获取行数和列数
            for row in range(1,rows):
                lst=[]
                # ncols是sheet的列数
                for col in range(0, cols):
                    # ctype为1是字符串，ctype为2是数值。
                    cell_type = sheet.cell(row, col).ctype
                    cell_value = sheet.cell_value(row, col)
                    # 去掉字符串首尾空格。excel会自动去掉整数和浮点数前后的空格。
                    if cell_type == 1:
                        lst.append(cell_value.strip())
                    # xlrd读取cell时1变成1.0。如果是数值，并且原本是整数，则返回整数。
                    elif cell_type == 2 and cell_value == round(cell_value):
                        lst.append(int(cell_value))
                    # 浮点数则不用额外处理
                    else:
                        lst.append(cell_value)
                lsts.append(lst)
        # 从excel sheet中获取@pytest.mark.parametrize()所需要的参数名和数据
        # 参数名格式化，格式为"a,b,c"
        param_name = ','.join(lsts[0])
        # 去掉参数名行，剩下的就是数据
        data = lsts[1:]
        return param_name, data



if __name__ == '__main__':
    a=Excel_rw(filename=api_case_path)
    #a.write_tiqu('a','b')
    #print(a.read_tiqu('zhuangtaima'))
    #a.read_api_cases()
    #print(a.sheet_to_pytestlist())
    print(a.get_col_value())